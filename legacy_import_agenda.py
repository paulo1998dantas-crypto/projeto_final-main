"""Idempotent Agenda R02 importer: planning, reprogramming and stage state by ITEM."""
import argparse, json, unicodedata
from datetime import datetime
from pathlib import Path
from uuid import uuid4
from openpyxl import load_workbook
from sqlalchemy import text
import database

EMPTY={'','-','0','N/A','NA','AG','?'}
STAGE={'VIDROS':'VIDROS','A/C':'A/C','PREP':'PREP','SERRA.':'SERRA','EXPE.':'EXPE','DESMONT':'DESMONT','ELÉTRICA':'ELÉTRICA','REVEST':'REVEST','BCO':'BCO','ACESSÓ.':'ACESSÓRIO','PLOTA.':'PLOTAGEM','LIBERA.':'LIBERAÇÃO'}
def clean(v):
    if v is None:return ''
    x=str(v).strip();return '' if x.upper() in EMPTY else x
def norm(v):return ''.join(c for c in unicodedata.normalize('NFKD',clean(v).upper()) if not unicodedata.combining(c))
def date(v):return v if isinstance(v,datetime) else None
def stage_status(v):
    x=norm(v)
    if x in {'S','SIM','OK'}:return 'CONCLUÍDA'
    if x in {'N/A','NA'}:return 'NÃO_APLICÁVEL'
    if x in {'N','NAO','NÃO'}:return 'PENDENTE'
    if x in {'P','PARCIAL'}:return 'EM_ANDAMENTO'
    return 'PENDENTE'
def main():
    p=argparse.ArgumentParser();p.add_argument('file');p.add_argument('--dry-run',action='store_true');p.add_argument('--report',default='agenda_import_report.json');a=p.parse_args();r={'dry_run':a.dry_run,'updated':0,'ignored':0,'unmatched':[],'schedules':0,'stages':0}
    wb=load_workbook(a.file,read_only=True,data_only=True,keep_links=False);ws=wb['AGENDA'];heads={clean(v).upper():i for i,v in enumerate(next(ws.iter_rows(min_row=5,max_row=5,values_only=True)))}
    with database.engine.begin() as c:
        for row,values in enumerate(ws.iter_rows(min_row=6,values_only=True),6):
            item=clean(values[heads['ITEM']]) if 'ITEM' in heads else ''
            if not item:continue
            key=f'{Path(a.file).name}:AGENDA:{item}:{row}'
            if c.execute(text('select 1 from erp_legacy_import_records where source_key=:k'),{'k':key}).first():r['ignored']+=1;continue
            work=c.execute(text('select id from erp_work_orders where numero_os=:item'),{'item':item}).first()
            if not work:r['unmatched'].append({'row':row,'item':item});continue
            work_id=str(work._mapping['id']); r['updated']+=1
            dates=[date(values[heads[x]]) for x in ('DATA 1','REPROGRAMA 1','REPROGRAMA 2') if x in heads and date(values[heads[x]])]
            if not a.dry_run:
                previous=None
                for index,new in enumerate(dates):
                    c.execute(text('insert into erp_work_order_schedules(id,work_order_id,data_anterior,nova_data,motivo,usuario,vigente) values(:id,:work,:old,:new,:reason,\'IMPORTADOR\',false)'),{'id':str(uuid4()),'work':work_id,'old':previous,'new':new,'reason':f'Agenda R02 histórico {index+1}'})
                    previous=new
                if previous:
                    c.execute(text('update erp_work_orders set data_comercial_prevista=:date,updated_at=now() where id=:id'),{'date':previous,'id':work_id})
                    c.execute(text('update erp_work_order_schedules set vigente=true where work_order_id=:id and nova_data=:date'),{'id':work_id,'date':previous})
                for label,code in STAGE.items():
                    if label not in heads:continue
                    c.execute(text("insert into erp_work_order_stages(id,work_order_id,stage_code,aplicavel,status,ordem,observacoes) values(:id,:work,:code,true,:status,:order,'Importado da Agenda R02') on conflict(work_order_id,stage_code) do update set status=excluded.status,observacoes=excluded.observacoes"),{'id':str(uuid4()),'work':work_id,'code':code,'status':stage_status(values[heads[label]]),'order':list(STAGE).index(label)+1})
                c.execute(text("insert into erp_legacy_import_records(source_key,source_file,source_sheet,source_item,entity_type,entity_id,payload) values(:k,:file,'AGENDA',:item,'WORK_ORDER',:id,cast(:payload as jsonb))"),{'k':key,'file':Path(a.file).name,'item':item,'id':work_id,'payload':json.dumps({'sequenciamento':clean(values[heads['SEQUENCIAMENTO']]) if 'SEQUENCIAMENTO' in heads else '', 'pedido_compras':clean(values[heads['PEDIDO DE COMPRAS']]) if 'PEDIDO DE COMPRAS' in heads else ''})})
            r['schedules']+=len(dates);r['stages']+=len(STAGE)
    wb.close();Path(a.report).write_text(json.dumps(r,ensure_ascii=False,indent=2,default=str),encoding='utf8');print(json.dumps(r,ensure_ascii=False,default=str))
if __name__=='__main__':main()
