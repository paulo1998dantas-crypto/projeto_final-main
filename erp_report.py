"""Relatório diário consolidado de Controle de Produção e Agenda do MES."""
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from erp_service import (
    LEAD_TIME_DAYS,
    STAGES,
    productive_cycle_window,
    stage_input_code,
    work_order_is_archived,
    work_order_situation,
)


CORE_HEADERS = [
    "ITEM", "Nº PROPOSTA", "SITUAÇÃO", "DATA ENTRADA", "DATA APROV. PV",
    "DATA A CONSIDERAR", "VENDEDOR", "MERCADO", "CLIENTE", "MUNICÍPIO", "UF",
    "MARCA - MODELO - VERSÃO", "MMV", "CHASSI", "MODELO", "TIPO DE VEÍCULO",
    "LINHA", "TRANSFORMAÇÃO", "COD. BCO", "CJ. BCO", "ACESSIBILIDADE",
    "LOTAÇÃO", "A/C", "TIPO AR", "AR QUENTE", "ACESSÓRIO", "PLOTAGEM",
    "DATA COMERCIAL", "INÍCIO REAL DE PRODUÇÃO", "TÉRMINO PRODUÇÃO",
    "DIAS PRODUÇÃO", "DATA SAÍDA",
    "ATRASO?", "INFO", "CHASSI 2", "AVARIAS", "ARQUIVADO",
]
STAGE_HEADERS = [
    "VIDROS", "A/C ", "PREP", "SERRA.", "EXPE.", "DESMONT", "ELÉTRICA",
    "REVEST", "BCO", "ACESSÓ.", "PLOTA.", "LIBERA.",
]
STAGE_HEADER_BY_CODE = dict(zip((item[0] for item in STAGES), STAGE_HEADERS))
CONTROL_HEADERS = [
    "B.O.", "OBSERVAÇÕES CONTROLE PRODUÇÃO", "OBSERVAÇÕES GERAIS",
    "SEQUENCIAMENTO", "PEDIDO DE COMPRAS", "Nº SEQUENCIA",
]

_STAGE_CODE_BY_TOKEN = {
    "VIDROS": "VIDROS",
    "AC": "A/C",
    "PREP": "PREP",
    "PREPARACAO": "PREP",
    "SERRA": "SERRA",
    "EXPE": "EXPE",
    "EXPEDICAO": "EXPE",
    "DESMONT": "DESMONT",
    "DESMONTAGEM": "DESMONT",
    "ELETRICA": "ELÉTRICA",
    "REVEST": "REVEST",
    "REVESTIMENTO": "REVEST",
    "BCO": "BCO",
    "BANCO": "BCO",
    "ACESSORIO": "ACESSÓRIO",
    "PLOTAGEM": "PLOTAGEM",
    "PLOTA": "PLOTAGEM",
    "LIBERACAO": "LIBERAÇÃO",
    "LIBERA": "LIBERAÇÃO",
}


def _canonical_stage_code(value):
    """Unifica nomes atuais e legados sem alterar os apontamentos persistidos."""
    raw = str(value or "").strip()
    ascii_value = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode()
    token = re.sub(r"[^A-Z0-9]", "", ascii_value.upper())
    return _STAGE_CODE_BY_TOKEN.get(token, raw)


def _canonical_stage_map(stage_map):
    canonical = {}
    for raw_code, stage in (stage_map or {}).items():
        code = _canonical_stage_code((stage or {}).get("stage_code") or raw_code)
        canonical[code] = stage
    return canonical


def _has_deadline_flow(situation):
    """Prazo comercial só existe depois que a O.S. entrou no fluxo produtivo."""
    ascii_value = unicodedata.normalize("NFKD", str(situation or "")).encode(
        "ascii", "ignore"
    ).decode().upper()
    normalized = re.sub(r"\s+", " ", ascii_value.replace("_", " ")).strip()
    return (
        normalized in {"EM PATIO", "EM PRODUCAO"}
        or normalized.startswith("FINALIZADA")
        or normalized.startswith("ENTREGUE")
    )


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _excel_value(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _yes_no(value):
    normalized = str(value or "").strip().upper()
    return "SIM" if normalized in {"SIM", "S", "TRUE", "1"} else "NÃO"


def _date_to_consider(arrival, approval):
    values = [item for item in (_as_date(arrival), _as_date(approval)) if item]
    return max(values) if values else None


def _days_between(start, end):
    start_date, end_date = _as_date(start), _as_date(end)
    return (end_date - start_date).days if start_date and end_date else None


def _delay_label(planned, finished, status):
    normalized_status = str(status or "").strip().upper().replace(" ", "_")
    if normalized_status == "CANCELADA":
        return ""
    planned_date = _as_date(planned)
    if not planned_date:
        return ""
    terminal = (
        normalized_status in {"FINALIZADA", "ENTREGUE", "RETIRADA", "ARQUIVADA"}
        and bool(finished)
    )
    # CONCLUIDA pode representar apenas a conclusão técnica em Suprimentos.
    # Sem uma data produtiva/saída real ela continua sendo comparada com hoje.
    terminal = terminal or (normalized_status == "CONCLUIDA" and bool(finished))
    comparison = _as_date(finished) if terminal and finished else date.today()
    if not comparison:
        return ""
    difference = (comparison - planned_date).days
    if terminal:
        if difference <= 0:
            return "FINALIZADO DENTRO DO PRAZO"
        return f"FINALIZADO COM ATRASO DE {difference} DIA(S)"
    if difference > 0:
        return f"EM ATRASO DE {difference} DIA(S)"
    return f"FALTAM {abs(difference)} DIA(S) PARA ATRASAR"


def _commercial_deadline(row):
    """Prazo canônico de 30/45 dias, independente da programação vigente."""
    explicit = _as_date(row.get("data_comercial_calculada"))
    if explicit:
        return explicit
    base = _date_to_consider(row.get("data_chegada"), row.get("data_aprovacao"))
    if not base:
        return None
    line = str(row.get("linha") or "").strip().upper()
    return base + timedelta(days=LEAD_TIME_DAYS.get(line, 30))


def _situation(row):
    if row.get("report_source") == "VEHICLE_ENTRY":
        return "RETIRADA" if str(row.get("entry_status") or "").upper() == "RETIRADA" else "AGUARDANDO O.S."
    return work_order_situation(
        row.get("status") or row.get("entry_status"),
        row.get("tipo_servico"),
        row.get("stage_configuration_status"),
    )


def _sequence_week(value):
    current = _as_date(value)
    if not current:
        return ""
    iso = current.isocalendar()
    return f"{iso.year} - SEMANA {iso.week:02d}"


def _query_report_data(conn):
    work_orders = [
        dict(row._mapping) for row in conn.execute(text("""
            select
                w.*,e.item_number,e.data_chegada,e.status as entry_status,
                e.observacoes as entry_notes,e.avarias,e.modelo_veicular,
                v.chassi,v.marca,v.modelo,v.versao,v.mmv,
                seq.sequencia as sequencia_persistida,
                seq.semana_planejada as semana_planejada_persistida,
                coalesce(po.purchase_orders,'') as purchase_orders
            from erp_work_orders w
            join erp_vehicle_entries e on e.id=w.vehicle_entry_id
            join erp_vehicles v on v.id=e.vehicle_id
            left join erp_work_order_sequences seq on seq.work_order_id=w.id
            left join lateral (
                select string_agg(distinct p.numero_oc, ', ' order by p.numero_oc) as purchase_orders
                from erp_purchase_order_lines l
                join erp_purchase_orders p on p.id=l.purchase_order_id
                where l.work_order_id=w.id
            ) po on true
            where w.is_current=true
            order by e.item_number
        """))
    ]
    awaiting_entries = [
        dict(row._mapping) for row in conn.execute(text("""
            select
                e.id,e.item_number,e.status as entry_status,e.data_chegada,
                e.cliente_nome,e.observacoes as entry_notes,e.avarias,e.modelo_veicular,
                v.chassi,v.marca,v.modelo,v.versao,v.mmv
            from erp_vehicle_entries e
            join erp_vehicles v on v.id=e.vehicle_id
            left join erp_work_orders w
              on w.vehicle_entry_id=e.id
             and w.is_current=true
            where w.id is null
            order by e.item_number
        """))
    ]
    for entry in awaiting_entries:
        # A entrada não é uma O.S. e deve continuar sendo apenas uma fila do
        # PCP. O marcador permite que a exportação a represente sem inventar
        # número de O.S., programação ou apontamentos produtivos.
        entry.update({
            "report_source": "VEHICLE_ENTRY",
            "status": entry.get("entry_status") or "AGUARDANDO_O_S",
            "stage_configuration_status": "PENDENTE",
            "technical_status": "ABERTA",
            "purchase_orders": "",
        })

    report_rows = sorted(
        [*work_orders, *awaiting_entries],
        key=lambda row: (row.get("item_number") is None, row.get("item_number") or 0),
    )
    if not report_rows:
        return [], {}, {}, {}

    work_order_ids = [row["id"] for row in work_orders]
    stages = defaultdict(dict)
    schedules = defaultdict(list)
    observations = defaultdict(list)
    if work_order_ids:
        for result in conn.execute(text("""
            select * from erp_work_order_stages
            where work_order_id=any(:ids)
            order by work_order_id,ordem
        """), {"ids": work_order_ids}):
            row = dict(result._mapping)
            stages[row["work_order_id"]][_canonical_stage_code(row["stage_code"])] = row
        for result in conn.execute(text("""
            select * from erp_work_order_schedules
            where work_order_id=any(:ids)
            order by work_order_id,created_at,id
        """), {"ids": work_order_ids}):
            row = dict(result._mapping)
            schedules[row["work_order_id"]].append(row)
        for result in conn.execute(text("""
            select work_order_id,observacao
            from erp_work_order_status_history
            where work_order_id=any(:ids)
              and nullif(trim(coalesce(observacao,'')),'') is not null
            order by work_order_id,created_at
        """), {"ids": work_order_ids}):
            observations[result._mapping["work_order_id"]].append(result._mapping["observacao"])
    awaiting_entry_ids = [row["id"] for row in awaiting_entries]
    if awaiting_entry_ids:
        try:
            pre_os_ready = bool(conn.execute(text(
                "select to_regclass('public.erp_vehicle_entry_stages') is not null"
            )).scalar())
        except (AttributeError, TypeError):
            pre_os_ready = False
        if pre_os_ready:
            for result in conn.execute(text("""
                select * from erp_vehicle_entry_stages
                where vehicle_entry_id=any(:ids)
                order by vehicle_entry_id,ordem
            """), {"ids": awaiting_entry_ids}):
                stage = dict(result._mapping)
                stages[stage["vehicle_entry_id"]][
                    _canonical_stage_code(stage["stage_code"])
                ] = stage
    return report_rows, stages, schedules, observations


def _report_row(row, stage_map, schedule_rows, status_notes, max_schedules):
    stage_map = _canonical_stage_map(stage_map)
    stage_notes = []
    stage_values = {}
    for code, _, _ in STAGES:
        stage = stage_map.get(code)
        if not stage:
            stage_value = (
                "?"
                if row.get("report_source") != "VEHICLE_ENTRY"
                and str(row.get("stage_configuration_status") or "").strip().upper() == "PENDENTE"
                else ""
            )
        elif row.get("report_source") == "VEHICLE_ENTRY" and not stage.get("parametrizado"):
            stage_value = ""
        else:
            stage_value = stage_input_code(stage)
        stage_values[STAGE_HEADER_BY_CODE[code]] = stage_value
        if stage and str(stage.get("observacoes") or "").strip():
            stage_notes.append(f"[{code}] {stage['observacoes']}")

    data_considerar = _date_to_consider(row.get("data_chegada"), row.get("data_aprovacao"))
    cycle_start, cycle_end = productive_cycle_window(row, list(stage_map.values()))
    cancelled = str(row.get("status") or "").strip().upper() == "CANCELADA"
    production_end = row.get("termino_producao") or (None if cancelled else cycle_end)
    end_reference = production_end or row.get("data_entrega")
    general_notes = [
        value for value in [row.get("entry_notes"), *status_notes]
        if str(value or "").strip()
    ]
    vehicle_description = " ".join(
        str(value).strip() for value in (row.get("marca"), row.get("modelo"), row.get("versao"))
        if str(value or "").strip()
    )
    current_schedule = next(
        (item for item in reversed(schedule_rows) if item.get("vigente")),
        schedule_rows[-1] if schedule_rows else None,
    )
    current_planned_date = (
        current_schedule.get("nova_data")
        if current_schedule
        else row.get("data_comercial_prevista")
    )
    situation = _situation(row)
    commercial_deadline = _commercial_deadline(row) if _has_deadline_flow(situation) else None
    purchase_order_references = list(dict.fromkeys(
        value.strip()
        for value in (
            str(row.get("purchase_orders") or ""),
            str(row.get("pedido_compras_legacy") or ""),
        )
        if value.strip()
    ))
    production_notes = list(dict.fromkeys(
        value.strip()
        for value in (
            str(row.get("observacoes_controle_producao") or ""),
            " | ".join(stage_notes),
        )
        if value.strip()
    ))
    general_notes = list(dict.fromkeys([
        *general_notes,
        str(row.get("observacoes_gerais") or "").strip(),
    ]))
    general_notes = [value for value in general_notes if value]
    values = {
        "ITEM": f"JI - {row['item_number']}",
        "Nº PROPOSTA": row.get("proposta_numero") or "",
        "SITUAÇÃO": situation,
        "DATA ENTRADA": row.get("data_chegada"),
        "DATA APROV. PV": row.get("data_aprovacao"),
        "DATA A CONSIDERAR": data_considerar,
        "VENDEDOR": row.get("vendedor") or "",
        "MERCADO": row.get("mercado") or "",
        "CLIENTE": row.get("cliente_nome") or "",
        "MUNICÍPIO": row.get("municipio") or "",
        "UF": row.get("uf") or "",
        "MARCA - MODELO - VERSÃO": vehicle_description,
        "MMV": row.get("mmv") or "",
        "CHASSI": row.get("chassi") or "",
        "MODELO": row.get("modelo_veicular") or "",
        "TIPO DE VEÍCULO": row.get("tipo_veiculo") or "",
        "LINHA": row.get("linha") or "",
        "TRANSFORMAÇÃO": row.get("transformacao") or "",
        "COD. BCO": row.get("codigo_banco") or "",
        "CJ. BCO": row.get("conjunto_bancos") or "",
        "ACESSIBILIDADE": row.get("acessibilidade") or "",
        "LOTAÇÃO": row.get("lotacao") or "",
        "A/C": row.get("ar_condicionado") or "",
        "TIPO AR": row.get("tipo_sistema_ar") or "",
        "AR QUENTE": row.get("ar_quente") or "",
        "ACESSÓRIO": row.get("acessorio") or "",
        "PLOTAGEM": row.get("plotagem") or "",
        "DATA COMERCIAL": commercial_deadline,
        "INÍCIO REAL DE PRODUÇÃO": cycle_start,
        "TÉRMINO PRODUÇÃO": production_end,
        "DIAS PRODUÇÃO": _days_between(row.get("data_aprovacao"), production_end),
        "DATA SAÍDA": row.get("data_entrega"),
        "ATRASO?": _delay_label(commercial_deadline, end_reference, row.get("status")),
        "INFO": row.get("info") or row.get("entry_notes") or "",
        "CHASSI 2": str(row.get("chassi") or "")[-8:],
        "AVARIAS": _yes_no(row.get("avarias")),
        # Arquivamento acompanha o encerramento produtivo/entrega do veículo.
        # A conclusão técnica de Suprimentos continua independente desse estado.
        "ARQUIVADO": "SIM" if work_order_is_archived(
            row.get("status"), row.get("technical_previous_status")
        ) else "NÃO",
        **stage_values,
        "B.O.": row.get("bo") or "",
        "OBSERVAÇÕES CONTROLE PRODUÇÃO": " | ".join(production_notes),
        "OBSERVAÇÕES GERAIS": " | ".join(dict.fromkeys(general_notes)),
        "SEQUENCIAMENTO": (
            row.get("semana_planejada_persistida")
            or row.get("sequenciamento_legacy")
            or _sequence_week(current_planned_date)
        ),
        "PEDIDO DE COMPRAS": " | ".join(purchase_order_references),
        "Nº SEQUENCIA": row.get("sequencia_persistida") or row.get("numero_sequencia_legacy") or "",
    }
    schedule_dates = [item.get("nova_data") for item in schedule_rows]
    values["DATA 1"] = schedule_dates[0] if schedule_dates else row.get("data_comercial_prevista")
    values["REPROGRAMA 1"] = current_planned_date
    return values


def _apply_sheet_style(sheet, headers, row_count):
    navy = PatternFill("solid", fgColor="123D6A")
    light_blue = PatternFill("solid", fgColor="DCEAF7")
    white_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="DCE4ED")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(row_count + 1, 2)}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 34
    for cell in sheet[1]:
        cell.fill = navy
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in sheet.iter_rows(min_row=2, max_row=row_count + 1):
        sheet.row_dimensions[row[0].row].height = 36
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    date_headers = {
        "DATA ENTRADA", "DATA APROV. PV", "DATA A CONSIDERAR", "DATA COMERCIAL",
        "DATA SAÍDA", "DATA 1",
    }
    date_headers.update(header for header in headers if header.startswith("REPROGRAMA "))
    datetime_headers = {"INÍCIO REAL DE PRODUÇÃO", "TÉRMINO PRODUÇÃO"}
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        if header in date_headers:
            for cell in sheet[letter][1:]:
                cell.number_format = "dd/mm/yyyy"
        elif header in datetime_headers:
            for cell in sheet[letter][1:]:
                cell.number_format = "dd/mm/yyyy hh:mm"
        width = 13
        if header in {"CHASSI", "MARCA - MODELO - VERSÃO", "TRANSFORMAÇÃO", "CJ. BCO"}:
            width = 28
        elif header in {
            "INFO", "B.O.", "OBSERVAÇÕES CONTROLE PRODUÇÃO", "OBSERVAÇÕES GERAIS",
            "TIPO AR", "ATRASO?",
        }:
            width = 25
        elif header in STAGE_HEADERS:
            width = 11
        sheet.column_dimensions[letter].width = width
    if row_count and "SITUAÇÃO" in headers:
        status_column = headers.index("SITUAÇÃO") + 1
        status_letter = get_column_letter(status_column)
        sheet.conditional_formatting.add(
            f"{status_letter}2:{status_letter}{row_count + 1}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("PARAMETRIZAÇÃO",{status_letter}2))'],
                fill=light_blue,
            ),
        )
        for header in (item for item in STAGE_HEADERS if item in headers):
            letter = get_column_letter(headers.index(header) + 1)
            sheet.conditional_formatting.add(
                f"{letter}2:{letter}{row_count + 1}",
                FormulaRule(formula=[f'{letter}2="?"'], fill=light_blue),
            )


def build_work_order_report(conn):
    work_orders, stages, schedules, observations = _query_report_data(conn)
    max_schedules = max((len(schedules[row["id"]]) for row in work_orders), default=0)
    max_schedules = max(max_schedules, 1)
    # A planilha principal é uma fotografia: primeira promessa e data vigente.
    # O histórico ilimitado continua normalizado na aba própria.
    schedule_headers = ["DATA 1", "REPROGRAMA 1"]
    headers = CORE_HEADERS + STAGE_HEADERS + CONTROL_HEADERS + schedule_headers

    workbook = Workbook()
    report = workbook.active
    report.title = "CONTROLE PRODUÇÃO"
    for column, header in enumerate(headers, start=1):
        report.cell(row=1, column=column, value=header)
    for row_number, work in enumerate(work_orders, start=2):
        values = _report_row(
            work, stages[work["id"]], schedules[work["id"]],
            observations[work["id"]], max_schedules,
        )
        for column, header in enumerate(headers, start=1):
            report.cell(row=row_number, column=column, value=_excel_value(values.get(header)))
    _apply_sheet_style(report, headers, len(work_orders))

    history = workbook.create_sheet("HISTÓRICO REPROGRAMAÇÕES")
    history_headers = [
        "ITEM", "O.S.", "CHASSI", "DATA ANTERIOR", "NOVA DATA",
        "MOTIVO", "USUÁRIO", "DATA/HORA", "SITUAÇÃO DA DATA",
    ]
    for column, header in enumerate(history_headers, start=1):
        history.cell(row=1, column=column, value=header)
    history_row = 2
    for work in work_orders:
        for schedule in schedules[work["id"]]:
            values = [
                f"JI - {work['item_number']}", work.get("numero_os"), work.get("chassi"),
                schedule.get("data_anterior"), schedule.get("nova_data"),
                schedule.get("motivo"), schedule.get("usuario"),
                schedule.get("created_at"),
                "DATA VIGENTE" if schedule.get("vigente") else "DATA ALTERADA",
            ]
            for column, value in enumerate(values, start=1):
                history.cell(row=history_row, column=column, value=_excel_value(value))
            history_row += 1
    _apply_sheet_style(history, history_headers, history_row - 2)
    for header in ("DATA ANTERIOR", "NOVA DATA", "DATA/HORA"):
        letter = get_column_letter(history_headers.index(header) + 1)
        for cell in history[letter][1:]:
            cell.number_format = "dd/mm/yyyy hh:mm" if header == "DATA/HORA" else "dd/mm/yyyy"
    history.column_dimensions["C"].width = 24
    history.column_dimensions["F"].width = 40

    legend = workbook.create_sheet("LEGENDA")
    legend.sheet_view.showGridLines = False
    legend.append(["RELATÓRIO DIÁRIO DO MES", "Significado"])
    legend.append(["Gerado em", datetime.now()])
    legend.append(["(vazio)", "Etapa ainda não preenchida em entrada sem O.S. ou registro sem parametrização pendente"])
    legend.append(["?", "Aguardando parametrização da etapa após a abertura da O.S."])
    legend.append(["P", "Parcial / em andamento"])
    legend.append(["N", "Pendente"])
    legend.append(["S", "Etapa concluída"])
    legend.append(["N/A", "Etapa não aplicável"])
    legend.append(["DATA COMERCIAL", "Prazo padrão: data a considerar + 30 dias (LB/LAB) ou 45 dias (LE/LAE)"])
    legend.append(["INÍCIO REAL DE PRODUÇÃO", "Primeiro início efetivamente apontado em qualquer etapa produtiva"])
    legend.append(["TÉRMINO PRODUÇÃO", "Data/hora da finalização produtiva; não é a entrega do veículo"])
    legend.append(["DATA SAÍDA", "Data/hora real da entrega ou retirada do veículo"])
    legend.append(["ATRASO?", "Calculado contra a DATA COMERCIAL padrão de 30/45 dias"])
    legend.append(["DATA 1", "Primeira promessa de entrega registrada para a O.S."])
    legend.append(["REPROGRAMA 1", "Data de entrega vigente na data da exportação"])
    legend.append(["DATA VIGENTE", "Data de programação atualmente consolidada para a O.S."])
    legend.append(["DATA ALTERADA", "Data substituída por uma reprogramação e mantida somente no histórico"])
    legend.append([
        "ARQUIVADO = SIM",
        "O.S. finalizada ou entregue, incluindo transformação, pós-venda e outros serviços.",
    ])
    legend.append([
        "Nº SEQUENCIA",
        "Mantido vazio enquanto não houver campo estrutural próprio; nunca é inferido pela linha do Excel.",
    ])
    legend["B2"].number_format = "dd/mm/yyyy hh:mm"
    legend.column_dimensions["A"].width = 24
    legend.column_dimensions["B"].width = 90
    for cell in legend[1]:
        cell.fill = PatternFill("solid", fgColor="123D6A")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in legend.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, len(work_orders), max_schedules
