import asyncio
import inspect
import unittest
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import main


class _FakeResult:
    def __init__(self, rows):
        self._rows = rows

    def mappings(self):
        return iter(self._rows)


class _FakeConnection:
    def __init__(self, rows):
        self.rows = rows
        self.sql = ""

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def execute(self, statement):
        self.sql = str(statement)
        return _FakeResult(self.rows)


class _FakeEngine:
    def __init__(self, rows):
        self.connection = _FakeConnection(rows)

    def connect(self):
        return self.connection


class MesOperationalExportTests(unittest.TestCase):
    @staticmethod
    async def _response_bytes(response):
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk.encode() if isinstance(chunk, str) else chunk)
        return b"".join(chunks)

    def test_history_export_reads_shared_erp_events(self):
        engine = _FakeEngine([{
            "item_number": 3110,
            "numero_os": "3110",
            "chassi": "9V7VPFC38TA004249",
            "modelo": "CITROEN JUMPY FURGAO",
            "stage_code": "VIDROS",
            "action": "APONTAMENTO",
            "status_anterior": "PENDENTE",
            "novo_status": "CONCLUÍDA",
            "operador": "PAULO",
            "inicio": datetime(2026, 7, 30, 10, tzinfo=timezone.utc),
            "termino": datetime(2026, 7, 30, 11, tzinfo=timezone.utc),
            "localizacao": "Linha",
            "created_at": datetime(2026, 7, 30, 11, tzinfo=timezone.utc),
        }])

        with patch.object(main.database, "engine", engine):
            rows = main._erp_history_export_rows()

        self.assertEqual(rows[0]["ITEM"], 3110)
        self.assertEqual(rows[0]["STATUS"], "CONCLUÍDA")
        self.assertEqual(rows[0]["ORIGEM"], "ERP")
        self.assertIsNone(rows[0]["DATA"].tzinfo)
        self.assertIn("erp_work_order_stage_events", engine.connection.sql)
        self.assertIn("erp_work_orders", engine.connection.sql)

    def test_time_export_reads_current_shared_erp_stages(self):
        engine = _FakeEngine([{
            "item_number": 3110,
            "numero_os": "3110",
            "chassi": "9V7VPFC38TA004249",
            "modelo": "CITROEN JUMPY FURGAO",
            "stage_code": "VIDROS",
            "status": "EM_ANDAMENTO",
            "responsavel": "PAULO",
            "inicio": None,
            "termino": None,
            "localizacao": "Linha",
        }])

        with patch.object(main.database, "engine", engine):
            rows = main._erp_time_export_rows()

        self.assertEqual(rows[0]["ETAPA"], "VIDROS")
        self.assertEqual(rows[0]["STATUS"], "EM_ANDAMENTO")
        self.assertEqual(rows[0]["ORIGEM"], "ERP")
        self.assertIn("erp_work_order_stages", engine.connection.sql)
        self.assertIn("erp_vehicle_entries", engine.connection.sql)

    def test_both_routes_keep_session_authentication(self):
        for endpoint in (main.exportar, main.exportar_tempos):
            source = inspect.getsource(endpoint)
            self.assertIn("require_login(request, db)", source)
            self.assertIn('RedirectResponse(url="/login", status_code=303)', source)

    def test_zero_erp_events_export_headers_without_legacy_schema(self):
        class _Inspector:
            @staticmethod
            def has_table(name):
                return name == "erp_work_order_stage_events"

        with (
            patch.object(main, "require_login", return_value=object()),
            patch.object(main, "has_permission", return_value=True),
            patch.object(main, "erp_feature_enabled", return_value=True),
            patch.object(main, "inspect", return_value=_Inspector()),
            patch.object(main, "_erp_history_export_rows", return_value=[]),
            patch.object(main, "legacy_operational_schema_available", return_value=False) as legacy,
        ):
            response = asyncio.run(main.exportar(object(), object()))

        legacy.assert_not_called()
        payload = asyncio.run(self._response_bytes(response))
        workbook = load_workbook(BytesIO(payload), read_only=True)
        headers = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]
        workbook.close()
        self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(
            headers,
            [
                "ITEM", "O.S.", "CHASSI", "MODELO", "ETAPA", "AÇÃO",
                "STATUS ANTERIOR", "STATUS", "RESPONSAVEL", "INICIO",
                "TERMINO", "LOCALIZACAO", "DATA", "ORIGEM",
            ],
        )

    def test_zero_erp_stages_export_headers_without_legacy_schema(self):
        class _Inspector:
            @staticmethod
            def has_table(name):
                return name == "erp_work_order_stages"

        with (
            patch.object(main, "require_login", return_value=object()),
            patch.object(main, "has_permission", return_value=True),
            patch.object(main, "erp_feature_enabled", return_value=True),
            patch.object(main, "inspect", return_value=_Inspector()),
            patch.object(main, "_erp_time_export_rows", return_value=[]),
            patch.object(main, "legacy_operational_schema_available", return_value=False) as legacy,
        ):
            response = asyncio.run(main.exportar_tempos(object(), object()))

        legacy.assert_not_called()
        payload = asyncio.run(self._response_bytes(response))
        workbook = load_workbook(BytesIO(payload), read_only=True)
        headers = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]
        workbook.close()
        self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(
            headers,
            [
                "ITEM", "O.S.", "CHASSI", "MODELO", "ETAPA", "STATUS",
                "RESPONSAVEL", "INICIO", "TERMINO", "LOCALIZACAO", "ORIGEM",
            ],
        )

    def test_export_buttons_are_available_in_dashboard_and_work_order_history(self):
        template_dir = Path(__file__).with_name("templates")
        index = (template_dir / "index.html").read_text(encoding="utf-8")
        management = (template_dir / "gestao_os.html").read_text(encoding="utf-8")
        for route in ("/exportar_historico", "/exportar_tempos"):
            self.assertGreaterEqual(index.count(route), 2)
            self.assertIn(route, management)


if __name__ == "__main__":
    unittest.main()
