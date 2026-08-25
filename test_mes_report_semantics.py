from datetime import date, datetime, timedelta
from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from erp_report import (
    CONTROL_HEADERS,
    CORE_HEADERS,
    STAGE_HEADERS,
    _commercial_deadline,
    _canonical_stage_code,
    build_work_order_report,
    _delay_label,
    _query_report_data,
    _report_row,
    _sequence_week,
)
from erp_service import work_order_is_archived


class MesReportSemanticsTests(unittest.TestCase):
    def base_work_order(self):
        return {
            "id": "wo-1",
            "item_number": 3110,
            "numero_os": "3110",
            "status": "EM_PRODUÇÃO",
            "stage_configuration_status": "CONCLUIDA",
            "data_chegada": date(2026, 7, 24),
            "data_aprovacao": date(2026, 7, 24),
            "data_comercial_calculada": date(2026, 8, 23),
            "data_comercial_prevista": date(2026, 8, 21),
            "marca": "MERCEDES-BENZ",
            "modelo": "SPRINTER 417",
            "versao": "10,5 M³",
            "modelo_veicular": "PACK",
            "technical_status": "ABERTA",
            "purchase_orders": "",
            "pedido_compras_legacy": "PC 123",
            "numero_sequencia_legacy": "1",
            "sequenciamento_legacy": "",
            "observacoes_controle_producao": "",
            "observacoes_gerais": "",
        }

    def test_commercial_deadline_is_distinct_from_current_schedule(self):
        row = _report_row(
            self.base_work_order(),
            {},
            [{
                "nova_data": date(2026, 8, 21),
                "vigente": True,
            }],
            [],
            1,
        )

        self.assertEqual(row["DATA COMERCIAL"], date(2026, 8, 23))
        self.assertEqual(row["REPROGRAMA 1"], date(2026, 8, 21))
        self.assertEqual(row["PEDIDO DE COMPRAS"], "PC 123")
        self.assertEqual(row["MODELO"], "PACK")
        self.assertEqual(row["MARCA - MODELO - VERSÃO"], "MERCEDES-BENZ SPRINTER 417 10,5 M³")
        self.assertEqual(row["Nº SEQUENCIA"], "1")
        self.assertNotIn("DATA ENTREGA", CONTROL_HEADERS)

    def test_latest_vigente_schedule_wins_without_losing_history(self):
        row = _report_row(
            self.base_work_order(),
            {},
            [
                {"nova_data": date(2026, 8, 20), "vigente": False},
                {"nova_data": date(2026, 8, 22), "vigente": True},
            ],
            [],
            2,
        )

        self.assertEqual(row["DATA 1"], date(2026, 8, 20))
        self.assertEqual(row["REPROGRAMA 1"], date(2026, 8, 22))

    def test_sequence_week_uses_finalization_after_vehicle_is_finished(self):
        work = self.base_work_order()
        work.update({
            "status": "FINALIZADA",
            "termino_producao": datetime(2026, 8, 20, 16, 30),
            "semana_planejada_persistida": "99",
            "sequenciamento_legacy": "SEMANA ANTIGA",
        })

        row = _report_row(
            work,
            {},
            [{"nova_data": date(2026, 9, 15), "vigente": True}],
            [],
            1,
        )

        self.assertEqual(_sequence_week(date(2026, 8, 20)), 34)
        self.assertEqual(row["SEQUENCIAMENTO"], 34)

    def test_sequence_week_uses_current_planned_date_before_finalization(self):
        work = self.base_work_order()
        work.update({
            "status": "EM_PRODUÇÃO",
            "termino_producao": None,
            "semana_planejada_persistida": "99",
        })

        row = _report_row(
            work,
            {},
            [{"nova_data": date(2026, 9, 15), "vigente": True}],
            [],
            1,
        )

        self.assertEqual(row["SEQUENCIAMENTO"], 38)

    def test_card_notes_and_purchase_orders_are_consolidated_in_their_columns(self):
        work = self.base_work_order()
        work.update({
            "purchase_orders": "2724 | 2749",
            "pedido_compras_legacy": "PC LEGADO",
            "observacoes_controle_producao": "Observação legada",
            "observacoes_gerais": "Não deve sair",
            "entry_notes": "Também não deve sair em observações gerais",
        })
        stages = {
            "PREP": {
                "stage_code": "PREP", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA", "observacoes": "Preparação conferida",
            },
        }

        row = _report_row(
            work,
            stages,
            [],
            ["Pedido aguardando material", "Cliente confirmou a cor"],
            0,
        )

        self.assertEqual(
            row["OBSERVAÇÕES CONTROLE PRODUÇÃO"],
            "Observação legada | Pedido aguardando material | Cliente confirmou a cor | "
            "[PREP] Preparação conferida",
        )
        self.assertEqual(row["OBSERVAÇÕES GERAIS"], "")
        self.assertEqual(row["PEDIDO DE COMPRAS"], "2724 | 2749 | PC LEGADO")

    def test_multiple_individual_banks_remain_in_one_report_cell(self):
        work = self.base_work_order()
        work.update({
            "codigo_banco": "10200001 / 10200003 / 10200007",
            "conjunto_bancos": "BCO FIXO 3L / BCO RECLINÁVEL 3L / BCO FIXO 1L",
        })

        row = _report_row(work, {}, [], [], 0)

        self.assertEqual(
            row["COD. BCO"],
            "10200001 / 10200003 / 10200007",
        )
        self.assertEqual(
            row["CJ. BCO"],
            "BCO FIXO 3L / BCO RECLINÁVEL 3L / BCO FIXO 1L",
        )

    def test_standard_commercial_deadline_is_calculated_from_line(self):
        work = self.base_work_order()
        work["data_comercial_calculada"] = None
        work["data_comercial_prevista"] = date(2026, 12, 31)
        work["linha"] = "LE"
        self.assertEqual(_commercial_deadline(work), date(2026, 9, 7))
        work["linha"] = "LB"
        self.assertEqual(_commercial_deadline(work), date(2026, 8, 23))

    def test_real_cycle_start_and_production_end_are_adjacent_and_distinct_from_delivery(self):
        work = self.base_work_order()
        work.update({
            "status": "ENTREGUE",
            "termino_producao": "2026-08-20T16:30:00",
            "data_entrega": "2026-08-21T09:00:00",
        })
        stages = {
            "PREP": {
                "stage_code": "PREP", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA", "inicio": "2026-08-01T08:15:00",
                "termino": "2026-08-01T10:00:00",
            },
            "LIBERAÇÃO": {
                "stage_code": "LIBERAÇÃO", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA", "inicio": "2026-08-20T15:00:00",
                "termino": "2026-08-20T16:30:00",
            },
        }
        row = _report_row(work, stages, [], [], 1)
        self.assertEqual(row["INÍCIO REAL DE PRODUÇÃO"], "2026-08-01T08:15:00")
        self.assertEqual(row["TÉRMINO PRODUÇÃO"], "2026-08-20T16:30:00")
        self.assertEqual(row["DATA SAÍDA"], "2026-08-21T09:00:00")
        self.assertLess(
            CORE_HEADERS.index("INÍCIO REAL DE PRODUÇÃO"),
            CORE_HEADERS.index("TÉRMINO PRODUÇÃO"),
        )

    def test_post_sale_delivery_is_explicit_in_exported_situation(self):
        work = self.base_work_order()
        work.update({"status": "ENTREGUE", "tipo_servico": "PÓS-VENDA"})
        row = _report_row(work, {}, [], [], 0)
        self.assertEqual(row["SITUAÇÃO"], "ENTREGUE PÓS-VENDAS")

    def test_other_finalization_is_explicit_in_exported_situation(self):
        work = self.base_work_order()
        work.update({"status": "FINALIZADA", "tipo_servico": "INSTALAÇÃO_DE_ACESSÓRIO"})
        row = _report_row(work, {}, [], [], 0)
        self.assertEqual(row["SITUAÇÃO"], "FINALIZADA OUTROS")

    def test_archived_follows_operational_finalization_for_every_service_type(self):
        for status, service_type in (
            ("FINALIZADA", "TRANSFORMAÇÃO"),
            ("FINALIZADA", "PÓS-VENDA"),
            ("FINALIZADA", "INSTALAÇÃO_DE_ACESSÓRIO"),
            ("ENTREGUE", "TRANSFORMAÇÃO"),
            ("ENTREGUE", "PÓS-VENDA"),
            ("ENTREGUE", "OUTRO"),
            ("ARQUIVADA", "TRANSFORMAÇÃO"),
        ):
            with self.subTest(status=status, service_type=service_type):
                work = self.base_work_order()
                work.update({
                    "status": status,
                    "tipo_servico": service_type,
                    "technical_status": "ABERTA",
                })
                self.assertTrue(work_order_is_archived(status))
                self.assertEqual(_report_row(work, {}, [], [], 0)["ARQUIVADO"], "SIM")

        technically_closed = self.base_work_order()
        technically_closed.update({
            "status": "CONCLUIDA",
            "technical_status": "CONCLUIDA",
            "technical_previous_status": "ENTREGUE",
        })
        self.assertTrue(work_order_is_archived("CONCLUIDA", "ENTREGUE"))
        self.assertEqual(
            _report_row(technically_closed, {}, [], [], 0)["ARQUIVADO"], "SIM"
        )

    def test_archived_does_not_follow_technical_closure_or_withdrawal(self):
        for status in ("RASCUNHO", "AGUARDANDO_O_S", "ATIVA", "EM_PRODUÇÃO", "RETIRADA", "CANCELADA"):
            with self.subTest(status=status):
                work = self.base_work_order()
                work.update({"status": status, "technical_status": "CONCLUIDA"})
                self.assertFalse(work_order_is_archived(status))
                self.assertEqual(_report_row(work, {}, [], [], 0)["ARQUIVADO"], "NÃO")

    def test_cancelled_order_is_not_reported_as_delayed(self):
        self.assertEqual(
            _delay_label(date(2026, 8, 20), date(2026, 8, 10), "CANCELADA"),
            "",
        )

    def test_legacy_stage_codes_are_exported_in_their_canonical_columns(self):
        work = self.base_work_order()
        stages = {
            "AC": {
                "stage_code": "AC", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA",
            },
            "ELETRICA": {
                "stage_code": "ELETRICA", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA",
            },
            "ACESSORIO": {
                "stage_code": "ACESSORIO", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA",
            },
            "LIBERACAO": {
                "stage_code": "LIBERACAO", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA",
            },
        }

        row = _report_row(work, stages, [], [], 0)

        self.assertEqual(_canonical_stage_code("A/C"), "A/C")
        self.assertEqual(row["A/C "], "S")
        self.assertEqual(row["ELÉTRICA"], "S")
        self.assertEqual(row["ACESSÓ."], "S")
        self.assertEqual(row["LIBERA."], "S")

    def test_commercial_deadline_and_delay_are_blank_outside_productive_flow(self):
        for status, configuration in (
            ("AGUARDANDO_O_S", "PENDENTE"),
            ("RASCUNHO", "PENDENTE"),
            ("CANCELADA", "CONCLUIDA"),
            ("RETIRADA", "CONCLUIDA"),
        ):
            with self.subTest(status=status):
                work = self.base_work_order()
                work.update({
                    "status": status,
                    "stage_configuration_status": configuration,
                })
                row = _report_row(work, {}, [], [], 0)
                self.assertEqual(row["DATA COMERCIAL"], None)
                self.assertEqual(row["ATRASO?"], "")

    def test_commercial_deadline_remains_for_patio_production_and_terminal_flow(self):
        for status in ("ATIVA", "EM_PRODUÇÃO", "FINALIZADA", "ENTREGUE"):
            with self.subTest(status=status):
                work = self.base_work_order()
                work["status"] = status
                row = _report_row(work, {}, [], [], 0)
                self.assertEqual(row["DATA COMERCIAL"], date(2026, 8, 23))
                self.assertTrue(row["ATRASO?"])

    def test_vehicle_entry_without_work_order_is_exported_as_awaiting_os(self):
        entry = {
            "id": "entry-3112",
            "report_source": "VEHICLE_ENTRY",
            "status": "AGUARDANDO_O_S",
            "entry_status": "AGUARDANDO_O_S",
            "stage_configuration_status": "PENDENTE",
            "technical_status": "ABERTA",
            "item_number": 3112,
            "data_chegada": date(2026, 8, 5),
            "cliente_nome": "CLIENTE TESTE",
            "entry_notes": "Entrada registrada; aguardando definição da O.S.",
            "avarias": "NÃO",
            "chassi": "9V8VPFC3XTA008976",
            "marca": "MERCEDES-BENZ",
            "modelo": "SPRINTER 417",
            "versao": "FURGÃO",
            "modelo_veicular": "STANDART",
            "mmv": "",
            "purchase_orders": "",
        }

        row = _report_row(entry, {}, [], [], 1)

        self.assertEqual(row["ITEM"], "JI - 3112")
        self.assertEqual(row["SITUAÇÃO"], "AGUARDANDO O.S.")
        self.assertEqual(row["DATA A CONSIDERAR"], date(2026, 8, 5))
        self.assertEqual(row["CHASSI"], "9V8VPFC3XTA008976")
        self.assertEqual(row["MODELO"], "STANDART")
        self.assertEqual(row["REPROGRAMA 1"], None)
        self.assertTrue(all(row[header] == "" for header in STAGE_HEADERS))

    def test_pre_os_pointing_is_exported_but_unfilled_entry_stages_remain_blank(self):
        entry = {
            "id": "entry-3112", "report_source": "VEHICLE_ENTRY",
            "status": "AGUARDANDO_O_S", "entry_status": "AGUARDANDO_O_S",
            "stage_configuration_status": "PENDENTE", "technical_status": "ABERTA",
            "item_number": 3112, "data_chegada": date(2026, 8, 5),
            "purchase_orders": "",
        }
        row = _report_row(entry, {
            "PREP": {
                "stage_code": "PREP", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA",
            },
            "SERRA": {
                "stage_code": "SERRA", "parametrizado": False, "aplicavel": True,
                "status": "PENDENTE",
            },
        }, [], [], 1)
        self.assertEqual(row["PREP"], "S")
        self.assertEqual(row["SERRA."], "")

    def test_question_mark_is_reserved_for_open_os_awaiting_parameterization(self):
        work = self.base_work_order()
        work.update({"status": "RASCUNHO", "stage_configuration_status": "PENDENTE"})
        row = _report_row(work, {
            "PREP": {
                "stage_code": "PREP", "parametrizado": False, "aplicavel": True,
                "status": "PENDENTE",
            },
        }, [], [], 1)
        self.assertEqual(row["PREP"], "?")
        self.assertEqual(row["SERRA."], "?")

    def test_delay_text_uses_standard_deadline_language(self):
        self.assertEqual(
            _delay_label(date(2026, 8, 20), date(2026, 8, 22), "FINALIZADA"),
            "FINALIZADO COM ATRASO DE 2 DIA(S)",
        )
        self.assertEqual(
            _delay_label(date.today() + timedelta(days=3), None, "EM_PRODUÇÃO"),
            "FALTAM 3 DIA(S) PARA ATRASAR",
        )

    def test_technical_closure_without_production_end_is_not_reported_as_finished(self):
        planned = date.today() - timedelta(days=2)
        self.assertEqual(
            _delay_label(planned, None, "CONCLUIDA"),
            "EM ATRASO DE 2 DIA(S)",
        )
        self.assertEqual(
            _delay_label(planned, None, "FINALIZADA"),
            "EM ATRASO DE 2 DIA(S)",
        )

    def test_generated_workbook_is_a_snapshot_with_first_and_current_promises(self):
        work = self.base_work_order()
        work.update({
            "termino_producao": datetime(2026, 8, 20, 16, 30),
            "data_entrega": datetime(2026, 8, 21, 9, 0),
            "status": "ENTREGUE",
        })
        report_data = (
            [work],
            {"wo-1": {"PREP": {
                "stage_code": "PREP", "parametrizado": True, "aplicavel": True,
                "status": "CONCLUÍDA", "inicio": datetime(2026, 8, 1, 8, 15),
            }}},
            {"wo-1": [
                {"nova_data": date(2026, 8, 20), "vigente": False},
                {"nova_data": date(2026, 8, 22), "vigente": True},
            ]},
            {"wo-1": []},
        )
        with patch("erp_report._query_report_data", return_value=report_data):
            output, row_count, history_depth = build_work_order_report(None)
        workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
        sheet = workbook["CONTROLE PRODUÇÃO"]
        headers = [cell.value for cell in sheet[1]]
        values = {header: sheet.cell(2, index + 1).value for index, header in enumerate(headers)}

        self.assertEqual(row_count, 1)
        self.assertEqual(history_depth, 2)
        self.assertNotIn("DATA ENTREGA", headers)
        self.assertEqual(headers[-2:], ["DATA 1", "REPROGRAMA 1"])
        self.assertEqual(values["DATA 1"], datetime(2026, 8, 20))
        self.assertEqual(values["REPROGRAMA 1"], datetime(2026, 8, 22))
        self.assertEqual(values["INÍCIO REAL DE PRODUÇÃO"], datetime(2026, 8, 1, 8, 15))
        self.assertEqual(values["TÉRMINO PRODUÇÃO"], datetime(2026, 8, 20, 16, 30))
        self.assertEqual(values["DATA SAÍDA"], datetime(2026, 8, 21, 9, 0))

    def test_report_query_includes_vehicle_entries_without_work_order(self):
        class FakeRow:
            def __init__(self, value):
                self._mapping = value

        class FakeConnection:
            def __init__(self):
                self.statements = []

            def execute(self, statement, _params=None):
                sql = str(statement)
                self.statements.append(sql)
                if "where w.id is null" in sql:
                    return [FakeRow({
                        "id": "entry-3112",
                        "item_number": 3112,
                        "entry_status": "AGUARDANDO_O_S",
                        "data_chegada": date(2026, 8, 5),
                        "cliente_nome": "CLIENTE TESTE",
                        "entry_notes": "",
                        "avarias": "NÃO",
                        "chassi": "9V8VPFC3XTA008976",
                        "marca": "MERCEDES-BENZ",
                        "modelo": "SPRINTER 417",
                        "versao": "FURGÃO",
                        "mmv": "",
                    })]
                return []

        connection = FakeConnection()
        rows, stages, schedules, observations = _query_report_data(connection)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["report_source"], "VEHICLE_ENTRY")
        self.assertEqual(rows[0]["status"], "AGUARDANDO_O_S")
        self.assertEqual(dict(stages), {})
        self.assertEqual(dict(schedules), {})
        self.assertEqual(dict(observations), {})
        sql = "\n".join(connection.statements)
        self.assertIn("p.work_order_id=w.id", sql)
        self.assertIn("p.vehicle_entry_id=e.id", sql)
        self.assertIn("erp_vehicle_entry_notes", sql)


if __name__ == "__main__":
    unittest.main()
