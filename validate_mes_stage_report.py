"""Valida parametrização pré-ativação e relatório XLSX sem persistir dados de teste."""
import json
from datetime import date
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import text

import database
from erp_report import build_work_order_report
from erp_service import (
    STAGES, activate_work_order, configure_stages, create_entry,
    create_work_order, reschedule, technical_close_work_order,
    update_stage, work_order_detail,
)


def main():
    connection = database.engine.connect()
    transaction = connection.begin()
    try:
        actor = "validacao-parametrizacao"
        entry = create_entry(connection, {
            "chassi": f"PARAM{uuid4().hex[:12]}".upper(),
            "cliente_nome": "VALIDAÇÃO RELATÓRIO",
            "marca": "CITROËN",
            "modelo": "JUMPY",
            "versao": "FURGÃO",
            "modelo_veicular": "PACK",
            "avarias": "NÃO",
            "observacoes": "Entrada criada somente dentro de transação de teste.",
        }, actor)
        work = create_work_order(connection, entry["id"], {
            "tipo_servico": "TRANSFORMAÇÃO",
            "proposta_numero": "REL-001",
            "data_aprovacao": date(2026, 7, 29),
            "vendedor": "JI",
            "mercado": "VAREJO",
            "cliente_nome": "VALIDAÇÃO RELATÓRIO",
            "municipio": "CAXIAS DO SUL",
            "uf": "RS",
            "tipo_veiculo": "MICRO",
            "linha": "LB",
            "transformacao_codigo": "40340009",
            "transformacao": "JI CONFORT 417 10 M SELADO ESSENCIAL",
            "conjunto_bancos": "CJ TESTE",
            "ar_condicionado": "GE",
            "tipo_sistema_ar": "COMPLEMENTO",
            "ar_quente": "NÃO",
            "data_comercial_prevista": date(2026, 8, 28),
        }, actor)
        detail = work_order_detail(connection, work["id"])
        assert len(detail["stages"]) == 12
        assert {stage["input_code"] for stage in detail["stages"]} == {"?"}

        try:
            activate_work_order(connection, work["id"], actor)
            raise AssertionError("Ativação sem parametrização não foi bloqueada.")
        except ValueError as exc:
            assert "parametrização" in str(exc)

        partial = configure_stages(
            connection, work["id"],
            {"stages": {"VIDROS": "S", "A/C": "N/A"}, "complete": False},
            actor,
        )
        assert len(partial["pending_stages"]) == 10
        # LIBERAÇÃO is only the final production pointing.  All other
        # applicable stages must already be concluded or not applicable.
        choices = {code: "S" for code, _, _ in STAGES}
        choices.update({"A/C": "N/A", "PREP": "N", "LIBERAÇÃO": "N"})
        completed = configure_stages(
            connection, work["id"], {"stages": choices, "complete": True}, actor,
        )
        assert completed["complete"] is True
        activate_work_order(connection, work["id"], actor)

        # Qualquer etapa P/S inicia a produção; LIBERAÇÃO concluída fecha o
        # ciclo somente depois de todas as etapas aplicáveis.
        started = update_stage(
            connection, work["id"], "PREP",
            {"input_code": "P", "confirmed_status_change": True, "idempotency_key": str(uuid4())}, actor,
        )
        assert started["work_order_status"] == "EM_PRODUÇÃO"
        update_stage(
            connection, work["id"], "PREP",
            {"input_code": "S", "confirmed_status_change": True, "idempotency_key": str(uuid4())}, actor,
        )
        completed_cycle = update_stage(
            connection, work["id"], "LIBERAÇÃO",
            {"input_code": "S", "confirmed_status_change": True, "idempotency_key": str(uuid4())}, actor,
        )
        assert completed_cycle["work_order_status"] == "FINALIZADA"
        cycle_detail = work_order_detail(connection, work["id"])
        assert cycle_detail["work_order"]["inicio_ciclo_produtivo"] is not None
        assert cycle_detail["work_order"]["fim_ciclo_produtivo"] is not None

        reschedule(connection, work["id"], date(2026, 9, 2), "Primeiro ajuste", actor)
        reschedule(connection, work["id"], date(2026, 9, 5), "Segundo ajuste", actor)
        technical_close_work_order(
            connection,
            work["id"],
            actor,
            "Conclusão técnica usada na validação do relatório.",
        )
        output, row_count, schedule_count = build_work_order_report(connection)
        workbook = load_workbook(output, data_only=False)
        report = workbook["CONTROLE PRODUÇÃO"]
        headers = [cell.value for cell in report[1]]
        assert row_count >= 1
        assert schedule_count >= 3
        assert "DATA 1" in headers and "REPROGRAMA 1" in headers
        assert "REPROGRAMA 2" not in headers
        assert "DATA ENTREGA" not in headers
        assert "INÍCIO REAL DE PRODUÇÃO" in headers
        assert "TÉRMINO PRODUÇÃO" in headers
        assert "HISTÓRICO REPROGRAMAÇÕES" in workbook.sheetnames
        assert "LEGENDA" in workbook.sheetnames
        item_column = headers.index("ITEM") + 1
        matching_rows = [
            row for row in range(2, report.max_row + 1)
            if report.cell(row, item_column).value == f"JI - {entry['item_number']}"
        ]
        assert len(matching_rows) == 1
        row = matching_rows[0]
        assert report.cell(row, headers.index("VIDROS") + 1).value == "S"
        assert report.cell(row, headers.index("A/C ") + 1).value == "N/A"
        assert report.cell(row, headers.index("DESMONT") + 1).value == "P"
        assert report.cell(row, headers.index("AVARIAS") + 1).value == "NÃO"
        assert report.cell(row, headers.index("ARQUIVADO") + 1).value == "SIM"
        schedule_history = workbook["HISTÓRICO REPROGRAMAÇÕES"]
        schedule_history_headers = [cell.value for cell in schedule_history[1]]
        date_status_column = schedule_history_headers.index("SITUAÇÃO DA DATA") + 1
        date_statuses = [
            schedule_history.cell(row_number, date_status_column).value
            for row_number in range(2, schedule_history.max_row + 1)
            if schedule_history.cell(row_number, 1).value == f"JI - {entry['item_number']}"
        ]
        assert date_statuses.count("DATA VIGENTE") == 1
        assert date_statuses.count("DATA ALTERADA") == 2
        status = connection.execute(text(
            "select status from erp_work_orders where id=:id"
        ), {"id": work["id"]}).scalar_one()
        assert status == "CONCLUIDA"
        print(json.dumps({
            "status": "PASS",
            "item": entry["item_number"],
            "stages": len(detail["stages"]),
            "pending_after_partial": len(partial["pending_stages"]),
            "stage_codes": {"VIDROS": "S", "A/C": "N/A", "DESMONT": "P"},
            "lifecycle": {"start": started["work_order_status"], "end": completed_cycle["work_order_status"]},
            "report_rows": row_count,
            "schedule_history_depth": schedule_count,
            "sheets": workbook.sheetnames,
            "transaction": "ROLLBACK",
        }, ensure_ascii=False, indent=2))
    finally:
        transaction.rollback()
        connection.close()


if __name__ == "__main__":
    main()
