"""Idempotent legacy importer for R08 production control and R02 Agenda.

Usage: DATABASE_URL=... python legacy_import.py --production FILE --agenda FILE --dry-run
Remove --dry-run only after reviewing the JSON report.
"""
import argparse, json, re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import text
import database, erp_service

EMPTY = {'', '-', '0', 'AG', 'N/A', 'NA', 'NONE', 'NAN'}
def clean(v):
    if v is None: return ''
    x=str(v).strip(); return '' if x.upper() in EMPTY else x
def date_value(v):
    if not v or clean(v)=='': return None
    if isinstance(v, datetime): return v
    return None
def header_map(ws, row): return {clean(ws.cell(row,c).value).upper(): c-1 for c in range(1,ws.max_column+1) if clean(ws.cell(row,c).value)}
def cell(values, headers, *names):
    for name in names:
        if name.upper() in headers: return values[headers[name.upper()]]
    return None
def status_map(value):
    raw=clean(value).upper()
    if any(x in raw for x in ('ENTREG','RETIR')): return 'ENTREGUE'
    if any(x in raw for x in ('FINAL','CONCLU')): return 'FINALIZADA'
    if any(x in raw for x in ('CANCEL','ARQUIV')): return 'ARQUIVADA'
    if raw: return 'RASCUNHO'
    return 'RASCUNHO'

def import_production(path, dry_run=True):
    report={'file':str(path),'sheet':'CONTROLE DE PRODUÇÃO','dry_run':dry_run,'inserted':0,'ignored':0,'rejected':[]}
    wb=load_workbook(path,read_only=True,data_only=True,keep_links=False)
    ws=wb['CONTROLE DE PRODUÇÃO']; headers=header_map(ws,3)
    with database.engine.begin() as conn:
        for r, values in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            item=clean(cell(values,headers,'ITEM')); chassi=clean(cell(values,headers,'CHASSI'))
            if not item and not chassi: continue
            source=f'{Path(path).name}:CONTROLE DE PRODUÇÃO:{item or r}'
            exists=conn.execute(text('select 1 from erp_legacy_import_records where source_key=:key'),{'key':source}).first()
            if exists: report['ignored']+=1; continue
            if not chassi: report['rejected'].append({'row':r,'item':item,'error':'chassi ausente'}); continue
            entry={'chassi':chassi,'mmv':clean(cell(values,headers,'MMV')),'modelo':clean(cell(values,headers,'MODELO','MARCA - MODELO - VERSÃO')),'cliente_nome':clean(cell(values,headers,'CLIENTE')),'data_chegada':date_value(cell(values,headers,'DATA ENTRADA')) or datetime.utcnow(),'origem':'LEGACY_R08','observacoes':clean(cell(values,headers,'INFORMAÇÕES')),'avarias':clean(cell(values,headers,'AVARIAS'))}
            work={'numero_os':item,'proposta_numero':clean(cell(values,headers,'Nº PROPOSTA')),'data_aprovacao':date_value(cell(values,headers,'DATA APROV. PV')),'vendedor':clean(cell(values,headers,'VENDEDOR')),'mercado':clean(cell(values,headers,'MERCADO')),'cliente_nome':entry['cliente_nome'],'municipio':clean(cell(values,headers,'MUNICÍPIO')),'uf':clean(cell(values,headers,'UF')),'tipo_veiculo':clean(cell(values,headers,'TIPO DE VEÍCULO')),'linha':clean(cell(values,headers,'LINHA')),'transformacao':clean(cell(values,headers,'TRANSFORMAÇÃO')),'codigo_banco':clean(cell(values,headers,'COD. BCO')),'conjunto_bancos':clean(cell(values,headers,'CJ. BCO')),'acessibilidade':clean(cell(values,headers,'ACESSIBILIDADE')),'lotacao':clean(cell(values,headers,'LOTAÇÃO')),'ar_condicionado':clean(cell(values,headers,'A/C')),'tipo_sistema_ar':clean(cell(values,headers,'TIPO AR')),'ar_quente':clean(cell(values,headers,'AR QUENTE')),'acessorio':clean(cell(values,headers,'ACESSÓRIO')),'plotagem':clean(cell(values,headers,'PLOTAGEM')),'data_comercial_prevista':date_value(cell(values,headers,'DATA COMERCIAL'))}
            if dry_run: report['inserted']+=1; continue
            created=erp_service.create_entry(conn,entry,'IMPORTADOR'); wo=erp_service.create_work_order(conn,created['id'],work,'IMPORTADOR')
            target=status_map(cell(values,headers,'SITUAÇÃO'))
            if target in {'FINALIZADA','ENTREGUE'}: erp_service.finalize(conn,wo['id'],'IMPORTADOR',target=='ENTREGUE','Importacao legado')
            conn.execute(text("insert into erp_legacy_import_records(source_key,source_file,source_sheet,source_item,entity_type,entity_id,payload) values(:key,:file,'CONTROLE DE PRODUÇÃO',:item,'WORK_ORDER',:id,cast(:payload as jsonb))"),{'key':source,'file':Path(path).name,'item':item,'id':wo['id'],'payload':json.dumps({'chassi':chassi,'status_legacy':clean(cell(values,headers,'SITUAÇÃO'))})})
            report['inserted']+=1
    wb.close(); return report

def main():
    p=argparse.ArgumentParser(); p.add_argument('--production',required=True); p.add_argument('--agenda'); p.add_argument('--dry-run',action='store_true'); p.add_argument('--report',default='legacy_import_report.json'); args=p.parse_args()
    result={'production':import_production(args.production,args.dry_run),'agenda':{'file':args.agenda,'status':'planned-import-not-required-for-R08-source'} if args.agenda else None}
    Path(args.report).write_text(json.dumps(result,ensure_ascii=False,indent=2,default=str),encoding='utf-8'); print(json.dumps(result,ensure_ascii=False,default=str))
if __name__=='__main__': main()
